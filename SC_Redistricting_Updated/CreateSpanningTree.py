# -*- coding: utf-8 -*-
"""
Created on Wed Jun 23 18:21:17 2021

@author: Blake Splitter and Amy Burton

This code should be designed to
take in a graph G=(V,E) and two adjacent districts. Then a subgraph imposed on
the nodes of those two districts will be constructed. Thereafter, we will create
a spanning tree on the resulting subgraph.
"""

###TO DO:
"#1. Change asterisk on 'with .... as cursor:' line to include actual field names --- DONE (Blake). "
"#2. Validate that districts are adjacent (Working code in rows 60-106) --- DONE (Amy)."
"#3. Figure out how to adjust script validation in ArcGIS Pro so that methodtype changes the rest of the input. --- CONCLUDED. Determined unnecessary (Blake)" 
#4. Edit "with .... as cursor:" line to include more robust SQL statement.
#5. Currently, this code assumes that our neighbor list include both sides of every edge. Should we build it so that it doesn't necessarily assume that?
"#6. Probably need to hard-code Wilson's algorithm --- DONE (Blake). Taken from Greg's code."
"#7. Need to determine how to split the spanning tree into two pieces and recover the two pieces. IDEA: maybe networkx has some functionality that will let it determine which edges are in which tree --- DONE (Blake)"
"#8. Allow Python users to specify input easily. --- DONE (Input can be specified if running this function from a different script by calling CreateSpanningTree.main(***insert args here***), Blake)"
"#9. Find an easier way to determine whether we're running from within Python or ArcGIS --- DONE. Now the code checks what sys.executable is, rather than examining argument input. (Blake)"
#10. Generalize field names for 'shapefile'
#11. Figure out how to cut edges and find subtree efficiently.
#12. Consider how to change ideal population number as algorithm progresses
#13. In nbrlist_fields, account for the 'nodes' column too and adjust the code so that shapes adjacent by a point are not adjacent
#14. Allow stateG to be a proper input parameter

import arcpy, os, sys
import random
seed = 17389
random.seed(seed)
#from random import randint
import networkx as nx
from openpyxl import load_workbook
#import numpy ### WE SHOULD TRY TO USE ONLY 'RANDOM' IF WE CAN

runspot = None #Global variable that will determine whether the code is started from ArcGIS or the Python console

def FindNamingFields(in_table):
    lstFields = arcpy.ListFields(in_table)
    namefields=[]
    distfields=[]
    for field in lstFields:
        #if field.name in  ["GEOID20", "Name20", "NAME20", "Name", "FID", "SOURCE_ID"]:
        if field.name in  ["src_GEOID20", "src_OBJECTID", "src_FID", "src_SOURCE_ID", "nbr_GEOID20", "nbr_OBJECTID", "nbr_FID", "nbr_SOURCE_ID"]:
            namefields.append(field.name)
        if field.name in ["src_dist", "nbr_dist"]:
            distfields.append(field.name)
        if field.name == "NODE_COUNT":
            distfields.append(field.name)
    nbrlist_fields = [item for sublist in [namefields, distfields] for item in sublist]
    return(namefields,distfields,nbrlist_fields)

def loopErasedWalk(graph, rng, v1 = None, v2 = None):
    '''Returns a loop-erased random walk between components v1 & v2'''
    if v1 is None:
        v1 = [rng.choice(sorted(list(graph.nodes)))]
    if v2 is None:
        v2 = [rng.choice(sorted(list(graph.nodes)))]

    v = rng.choice(sorted(v1))
    walk = [v]
    while v not in v2:
        v = rng.choice(sorted(list(graph.neighbors(v))))
        if v in walk:
            walk = walk[0:walk.index(v)]
        walk.append(v)

    return walk

def wilson(graph, rng):
    '''Returns a uniform spanning tree on G'''
    walk = loopErasedWalk(graph, rng)
    currentNodes = [n for n in walk]

    uniformTree = nx.Graph()
    for i in range(len(walk)-1):
        uniformTree.add_edge(walk[i], walk[i+1])

    treeNodes = set(uniformTree.nodes)
    neededNodes = set(graph.nodes) - treeNodes

    while neededNodes:
        v = rng.choice(sorted(list(neededNodes))) # sort for code repeatability
        walk = loopErasedWalk(graph, rng, v1 = [v], v2 = currentNodes)
        currentNodes += walk
        for i in range(len(walk)-1):
            uniformTree.add_edge(walk[i], walk[i+1])
        treeNodes = set(uniformTree.nodes)
        neededNodes = set(graph.nodes) - treeNodes

    return uniformTree

def FindEdgeCut(tree,tol,criteria):
    '''Input a tree graph, a percent tolerance, and a criteria for the graph. The function will remove a 
    random edge that splits the tree into two pieces such that each piece 
    has criteria (like population) within that percent tolerance. The variable 'tol' should be a positive real 
    number in (0,100]. The criteria should be string that labels an attribute of the nodes of the tree.'''
    if tol > 100 or tol <=0 or (isinstance(tol,float)==False and isinstance(tol,int)==False): 
        arcerror("tol must be a float or integer variable in the range (0,100].")
    if nx.is_tree(tree) == False:
        arcerror("The input graph must be a tree.")
    if isinstance(criteria,str)==False:
        arcerror("The criteria input should be a string.")
    tree_edge_list = list(tree.edges)
    random.shuffle(tree_edge_list) #Randomly shuffles the edges of T
    e=None
    TELL= len(tree_edge_list) #TELL = Tree Edge List Length
    
    for i in range(TELL):
        e = tree_edge_list[i] #Edge to delete
        tree.remove_edge(*e)
        #arcprint("The edges of T are now {0}. We just deleted {1}.",tree.edges,e)
        subgraphs = nx.connected_components(tree)
        subgraphs_lst = list(subgraphs)
        arcprint("The subgraph candidates are {0}.",subgraphs_lst)
        dist_crit1 = sum(value for key, value in nx.get_node_attributes(tree,criteria).items() if key in subgraphs_lst[0]) #Finds population sum for first district
        dist_crit2 = sum(value for key, value in nx.get_node_attributes(tree,criteria).items() if key in subgraphs_lst[1]) #Finds population sum for second district
        total_crit= dist_crit1+dist_crit2
        if abs(dist_crit1 - total_crit/2) > 0.01*tol*(total_crit/2):
            tree.add_edge(*e) #Adds the edge back to the tree if it didn't meet the tolerance
        else:
            arcprint("Criteria requirement was met. Removing edge {0}. Required {1} iteration(s).\nThe two subgraphs are {2}, with {3} of {4} and {5}, respectively.",e,i+1,subgraphs_lst,criteria,dist_crit1,dist_crit2)
            return(dist_crit1,dist_crit2,subgraphs_lst)
        if i==TELL-1:
            arcprint("No subgraphs with appropriate criteria requirements were found.\n")
            return(float('inf'), float('inf'),[]) 

def arcprint(message,*variables):
    '''Prints a message using arcpy.AddMessage() unless it can't; then it uses print. '''
    if runspot == "ArcGIS":
        arcpy.AddMessage(message.format(*variables))
    elif runspot == "console":
        newmessage=message
        j=0
        while j<len(variables): #This while loop puts the variable(s) in the correct spot(s) in the string
            newmessage = newmessage.replace("{"+str(j)+"}",str(variables[j])) #Replaces {i} with the ith variable
            j=j+1
        print(newmessage)
    else: 
        raise RuntimeError("No value for runspot has been assigned")

def arcerror(message,*variables):
    '''Prints an error message using arcpy.AddError() unless it can't; then it uses print. '''
    if runspot == "ArcGIS":
        arcpy.AddError(message.format(*variables))
    elif runspot == "console":
        newmessage=message
        j=0
        while j<len(variables): #This while loop puts the variable(s) in the correct spot(s) in the string
            newmessage = newmessage.replace("{"+str(j)+"}",str(variables[j])) #Replaces {i} with the ith variable
            j=j+1
        raise RuntimeError(newmessage)
    else: 
        raise RuntimeError("No value for runspot has been assigned")

#%% MAIN CODE STARTS HERE
def main(*args):
    
    global runspot #Allows runspot to be changed inside a function
    
    if sys.executable == r"C:\Program Files\ArcGIS\Pro\bin\ArcGISPro.exe": #Change this line if ArcGIS is located elsewhere
        runspot = "ArcGIS"
        if __name__ == "__main__":
            arcprint("We are running this from inside ArcGIS")
    else:
        runspot = "console"
        if __name__=="__main__":
            arcprint("We are running this from the python console")
    
    currentdir = os.getcwd()
    path = currentdir + "\\SC_Redistricting_Updated.gdb"
    arcpy.env.workspace = path
    
    arcpy.env.overwriteOutput = True
    
    try: #First attempts to take input from system arguments (Works for ArcGIS parameters, for instance)
        neighbor_list = sys.argv[1]
        dist1=int(sys.argv[2])
        dist2=int(sys.argv[3])
        shapefile=sys.argv[4]
        tol=float(sys.argv[5])
        ###NEED TO ADD stateG HERE SOMEHOW
    except IndexError: 
        try: #Second, tries to take input from explicit input into main()
            neighbor_list = args[0]
            dist1 = int(args[1])
            dist2 = int(args[2])
            shapefile = args[3]
            tol=float(args[4])
            stateG = args[5]
            arcprint("Using input from another file")
        except IndexError: #Finally, manually assigns input values if they aren't provided
            neighbor_list=path+"\\tl_2020_45_county20_SpatiallyConstrainedMultivariateClustering1_neighbor_list_shapes"
            dist1=6
            #dist1=randint(1,7) #Randomly selecting districts
            dist2=4
            #dist2=randint(1,7) #Randonly selecting districts
            shapefile=path+"\\tl_2020_45_county20_SpatiallyConstrainedMultivariateClustering1"
            tol=30
            arcprint("We are using default input choices")
    
    excel_NL = "excel_NL.xlsx"
    arcpy.TableToExcel_conversion(neighbor_list, excel_NL)
    wb = load_workbook(filename = excel_NL)
    ws = wb.active
    NL_cols = ws.max_column
    NL_rows = ws.max_row
    
    try: 
        stateG = stateG #Doesn't do anything, but if stateG doesn't exist, it will produce an UpboundLocalError
    except UnboundLocalError:
        #global stateG ## Maybe unnecessary?
        stateG = nx.Graph() #Creates an empty graph that will contain all adjacencies for the state
    
    ## The following lines start the structure of repeating iterations by currently keep the script from successfully running inside ArcGIS
    #StopCriterion = False #Currently runs 3 iterations of successfully finding adjacent districts then quits. 
    #IterationCount = 0
    #while StopCriterion == False:
    
    if dist1==dist2:
        arcerror("The districts must be different. Currently, dist1={0} and dist2={1}.",dist1,dist2) #Creates an error if the two district choices are the same
        #arcprint("The districts must be different. Currently, dist1={0} and dist2={1}.",dist1,dist2) #Instead of breaking the system, we return to picking new districts
        #arcprint("neighbor_list is {0} and has type {1}",  neighbor_list, type(neighbor_list))
        #continue
        
    #Blake's attempt to replace neighbor_list references with excel references
#    fieldexist = False
#    first_row = list(ws.rows)[0]
#    for cell in first_row:
#        if cell.value == "src_dist":
#            fieldexist=True
#            break
#    if fieldexist==False:
#        ws.cell(row=1,column=NL_cols+1).value = "src_dist"
#        ws.cell(row=1,column=NL_cols+2).value = "nbr_dist"
#        NL_cols = NL_cols+2
#        for i in range(NL_cols-1,NL_cols+1):
#            for j in range(2,NL_rows+1):
#                ws.cell(row=j,column=i).value = ws.cell(row=j,column=i-4).value
        
        
        
    
    fieldexist=False
    lstFields = arcpy.ListFields(neighbor_list)
    for field in lstFields:
        if field.name == "src_dist":
            fieldexist=True
            break
    if fieldexist==False:
        arcpy.AddField_management(neighbor_list, "src_dist", "SHORT", field_alias="Source District")
        arcpy.AddField_management(neighbor_list, "nbr_dist", "SHORT", field_alias="Neighbor District")
    lstFields = arcpy.ListFields(neighbor_list)
    orig_dist_names=[]
    for field in lstFields:
        if field.name in ["src_CLUSTER_ID", "src_Dist_Assgn", "nbr_CLUSTER_ID", "nbr_Dist_Assgn"]:
            orig_dist_names.append(field.name)
    odn=orig_dist_names #An alias
    
    if fieldexist==False:
        with arcpy.da.UpdateCursor(neighbor_list, [odn[0],odn[1],'src_dist', 'nbr_dist']) as cursor:
            for row in cursor:
                row[2]=row[0]
                row[3]=row[1]
                cursor.updateRow(row)
    
    [namefields,distfields,nbrlist_fields] = FindNamingFields(neighbor_list)
    nlf = nbrlist_fields #An alias
    #NFL = len(namefields) #NFL = Name Fields Length (How many fields name the polygons)
    #DFL = len(distfields) #DFL = District Fields Length (How many fields denote the district number)
    

    AdjFlag=0
    with arcpy.da.SearchCursor(neighbor_list, nlf, '''{}={} AND {}={} AND {}={}'''.format("src_dist",dist1,"nbr_dist",dist2,"NODE_COUNT",0)) as cursor:
        for row in cursor:
            AdjFlag+=1
            if AdjFlag>=1:
                arcprint("Blake's Code: Adjacency Established between districts {0} and {1} by units {2} and {3}", dist1, dist2, row[0],row[1])
                break

#    ## Where Amy's code edits start.
#    dist1_bdnds = [] #Creates empty list of boundary units for dist1
#    dist2_bdnds = [] #Creates empty list of boundary units for dist2
#    
#    #Fills list of boundary units for dist1 and dist2
#    with arcpy.da.SearchCursor(shapefile, ["OBJECTID", "Cluster_ID"], """{}={} AND ({}={} OR {}={})""".
#                           format("Boundary",1, "Cluster_ID", dist1,"Cluster_ID",dist2)) as cursor: #Limits search to rows containing units from dist1 and dist2
#        for row in cursor:
#            if row[1]==dist1: #If ClusterID==dist1 and Boundary unit is Yes
#                if dist1_bdnds.count(row[0])==0: #If we haven't already added the unit, add it to the list
#                    dist1_bdnds.append(row[0])
#            elif row[1]==dist2: #If ClusterID==dist2 and Boundary unit is Yes
#                if dist2_bdnds.count(row[0])==0:  #If we haven't already added the unit, add it to the list
#                    dist2_bdnds.append(row[0])
#    if len(dist1_bdnds)<=len(dist2_bdnds): #Determine the district with the fewest boundary units
#        pridist = dist1 #primary district
#        secdist = dist2 #secondary district
#    else:
#        pridist = dist2
#        secdist = dist1
#    
#    AdjFlag = False
#    if dist1==pridist:
#        dist_tuple = tuple(dist1_bdnds)
#    else: 
#        dist_tuple = tuple(dist2_bdnds)
#    #with arcpy.da.SearchCursor(neighbor_list, ["src_OBJECTID", "nbr_OBJECTID"], """({} IN {}) AND {}={}""".
#    #                               format("src_OBJECTID", dist_tuple,"nbr_CLUSTER_ID",secdist)) as cursor:
#    with arcpy.da.SearchCursor(neighbor_list, [nlf[0], nlf[1]], """({} IN {}) AND {}={}""".
#                                   format(nlf[0], dist_tuple,nlf[4],secdist)) as cursor:
#        for row in cursor:
#            AdjFlag = True
#            arcprint("Adjacency Established between districts {0} and {1} by units {2} and {3}", pridist, secdist, row[0],row[1])
#            break
    if AdjFlag==0: #Instead of breaking we return back to pick new districts
        #I've added back in the error -- Blake
        arcerror("Districts {0} and {1} are not adjacent.",dist1, dist2)
        #continue
    
    ## Where Amy's code edits end.
    
    ##This next section of code updates the district neighbor list at each iteration
    #dist_neighbor_list = shapefile + "\\dist_neighbor_list"
    
#    arcpy.PolygonNeighbors_analysis(shapefile, "dist_neighbor_list", "CLUSTER_ID",None,None,None,"KILOMETERS")
#    
#    AdjFlag = False
#    with arcpy.da.SearchCursor("dist_neighbor_list", ["src_CLUSTER_ID","nbr_CLUSTER_ID"], """{}={}""".format("src_CLUSTER_ID",dist1)) as cursor:
#        for row in cursor:
#            if row[0]==dist1 and row[1]==dist2:
#                AdjFlag=True
#                arcprint("Districts {0} and {1} are adjacent.",dist1,dist2)
#                break
#    if AdjFlag ==False:
#        arcerror("Districts {0} and {1} are not adjacent.",dist1,dist2)
    
    
    G = nx.Graph() #Creates an empty graph that will contain adjacencies for the two districts
    nodes = [] #Creates empty node list
    edges = [] #Creates empty edge list
    distnum = {} #Initializes a dictionary that will contain the district number for each polygon
    popnum = {} #Initializes a dictionary that will contain the population for each polygon
    
    """RIGHT NOW, THE CODE FINDS THE GRAPH ON EACH ITERATION FOR THE TWO DISTRICTS. I WOULD INSTEAD LIKE TO CREATE AN ENTIRE ADJACENCY GRAPH FOR THE WHOLE STATE AND ON EACH ITERATION, FIND A SUBGRAPH WITH ONLY THE TWO DISTRICTS"""
    
    
    if nx.is_empty(stateG)==True:
        with arcpy.da.SearchCursor(shapefile,["OBJECTID","SUM_Popula"]) as cursor:
            for row in cursor:
                popnum[row[0]] = row[1]
                stateG.add_node(row[0])
        with arcpy.da.SearchCursor(neighbor_list,nbrlist_fields) as cursor:
            for row in cursor:
                cursor.reset
                if list(stateG.edges).count([row[0],row[1]])==0 and list(stateG.edges).count([row[1],row[0]])==0:
                    #edges.append([row[0],row[1]])
                    stateG.add_edge(row[0],row[1])
                distnum[row[0]]=row[3] 
        nx.set_node_attributes(stateG,popnum,"Population")
        nx.set_node_attributes(stateG,distnum,"District Number")
    nodes_for_G = []
    if distnum != {}:
        for i in distnum:
            if distnum[i]==dist1 or distnum[i]==dist2:
                nodes_for_G.append(i)
    else:
        distdict = dict(stateG.nodes("District Number"))
        for v in distdict:
            if distdict[v] == dist1 or distdict[v] == dist2:
                nodes_for_G.append(v)
    G = stateG.subgraph(nodes_for_G) #Finds a subgraph containing all adjacencies for vertices in the two districts
    
    

    
    
# Following line requires two-sided neighbor relationship. Maybe fix later.
#    with arcpy.da.SearchCursor(neighbor_list, nbrlist_fields, """{}={} OR {}={}""".format(distfields[0], dist1,distfields[0],dist2)) as cursor:
#    for row in cursor:
#        if nodes.count(row[0])==0:
#            nodes.append(row[0]) # If the node is not in the nodes list, add it
#            G.add_node(row[0])
#        if edges.count([row[0],row[1]])==0 and edges.count([row[1],row[0]])==0 and (row[NFL]==dist1 or row[NFL]==dist2) and (row[NFL+1]==dist1 or row[NFL+1]==dist2):
#            edges.append([row[0],row[1]]) #If the edge is not in the edge list AND the both vertices are either dist1 or dist2, add the edge
#            G.add_edge(row[0],row[1])   
    
    arcprint("Edges of G are {0}",G.edges)
    arcprint("Vertices of G are {0}",G.nodes)
    
    T = wilson(G,random) #Creates a uniform random spanning tree for G using Wilson's algorithm
    arcprint("T edges are {0}",T.edges)
    if popnum != {}:   
        nx.set_node_attributes(T,popnum,"Population") 
    else:
        nx.set_node_attributes(T,dict(G.nodes("Population")),"Population")
        
    if distnum != {}:   
        nx.set_node_attributes(T,distnum,"District Number") 
    else:
        nx.set_node_attributes(T,dict(G.nodes("District Number")),"District Number")
        
    [dist1_pop, dist2_pop,subgraphs] = FindEdgeCut(T,tol,"Population") #Removes an edge from T so that the Population of each subgraph is within tolerance (tol)
    
    #This next section of code decides which subgraph should become district 1 and which should become district 2
    if dist1_pop!=float('inf') and dist2_pop!=float('inf'):
        s0d1count=0
        s0d2count=0
        s1d1count=0
        s1d2count=0
        for i in subgraphs[0]: 
            if stateG.nodes[i]["District Number"] == dist1:
                s0d1count +=1
            elif stateG.nodes[i]["District Number"] == dist2:
                s0d2count +=1
        for i in subgraphs[1]:
            if stateG.nodes[i]["District Number"] == dist1:
                s1d1count +=1
            elif stateG.nodes[i]["District Number"] == dist2:
                s1d2count +=1
        
        if s0d1count + s1d2count >= s0d2count + s1d1count:
            for i in subgraphs[0]:
                stateG.nodes[i]["District Number"] = dist1
            for i in subgraphs[1]:
                stateG.nodes[i]["District Number"] = dist2
            arcprint("Subgraph 0 is the new district {0} and subgraph 1 is the new district {1}",dist1,dist2)
            
        else:
            for i in subgraphs[0]:
                stateG.nodes[i]["District Number"] = dist2
            for i in subgraphs[1]:
                stateG.nodes[i]["District Number"] = dist1
            arcprint("Subgraph 0 is the new district {0} and subgraph 1 is the new district {1}",dist2,dist1)
        
        #Updates the neighbor list table after each iteration
#        with arcpy.da.UpdateCursor(neighbor_list,nbrlist_fields,'''{}={} OR {}={} OR {}={} OR {}={}'''.format(nlf[3], dist1, nlf[4], dist1, nlf[3], dist2, nlf[4], dist2)) as cursor:
#                for row in cursor:
#                    if row[0] in G.nodes:
#                        row[3] = G.nodes[row[0]]["District Number"]
#                    if row[1] in G.nodes:
#                        row[4] = G.nodes[row[1]]["District Number"]
#                    cursor.updateRow(row)
                        
                        
    if __name__ != "__main__":
        return(dist1_pop, dist2_pop,stateG,G,nlf)
    #    IterationCount += 1
    #    if IterationCount==1:
    #        StopCriterion = True

if __name__ == "__main__":
    main()
