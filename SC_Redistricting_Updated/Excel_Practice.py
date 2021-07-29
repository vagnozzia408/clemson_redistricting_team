# -*- coding: utf-8 -*-
"""
Created on Thu Jul 29 01:02:10 2021

@author: blake
"""

from openpyxl import load_workbook
import os
import arcpy
from itertools import product

currentdir = os.getcwd()
path = currentdir + "\\SC_Redistricting_Updated.gdb"
arcpy.env.workspace = path

neighbor_list=path+"\\tl_2020_45_county20_SpatiallyConstrainedMultivariateClustering1_neighbor_list_shapes"

excel_NL = "excel_NL.xlsx"
arcpy.TableToExcel_conversion(neighbor_list, excel_NL)

wb = load_workbook(filename = excel_NL)
ws = wb.active
NL_rows = ws.max_row
NL_cols = ws.max_column
print(ws.max_row)
print(ws.max_column)

first_row = list(ws.rows)[0]
first_col = list(ws.columns)[0]
for cell in first_row:
    if cell.value == "src_dist":
        print("Hey! src_dist is here!")

#data = [ws.cell(row=1,column=i).value for i in range(1,ws.max_column)]
data = [ws.cell(row=i,column=j).value for i,j in product(range(1,ws.max_row), range(1,ws.max_column))]

ws.cell(row=1,column=15).value = 1738


ws.rows
wb.save(excel_NL)